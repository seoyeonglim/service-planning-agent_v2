#!/usr/bin/env python3
"""내부 WBS 마크다운 → 엑셀 변환 (개요·주차체계·리소스판·WBS 간트·미결)

사용법:
    python3 .claude/scripts/wbs_to_excel.py docs/[프로젝트명]
    python3 .claude/scripts/wbs_to_excel.py docs/[프로젝트명] --out 경로.xlsx

스킬 13 형식의 WBS(`wbs/WBS_*.md`)를 읽어 시트 5개로 변환한다.
자동 생성물이므로 엑셀을 직접 고치지 말고 마크다운을 고친 뒤 재실행한다.
"""
import argparse
import re
import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# 라이프플래닛 그린 계열(.claude/assets/pdf 팔레트와 동일 계통)
GREEN = "00A862"
GREEN_DK = "007A47"
GREEN_LT = "E6F6EF"
GREY = "F2F2F2"
GREY_TX = "808080"
AMBER = "FFF4CE"
RED = "C00000"

THIN = Side(style="thin", color="D9D9D9")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
GATE_SIDE = Side(style="medium", color=RED)      # 공식 게이트 검수 주
TGT_SIDE = Side(style="dashed", color=GREEN_DK)  # 내부 목표 주

GATES, TARGETS = [], []
WEEK_START = {}   # 주차 → 시작일(MM.DD)

# 투입률 → 채움색
RATE_FILL = {
    100: "00A862",
    75: "4FC48D",
    50: "9BDCBF",
    25: "D6F0E4",
    0: "FFFFFF",
}

STAGE_FILL = {
    "M0": "FFE9C7",
    "P1": "D9E8FB",
    "P2": "E3DBF5",
    "P3-1": "FBE0E4",
    "P3-2": "FBD5DB",
    "오픈": "DFF3E6",
}


def md_tables(text):
    """마크다운에서 (헤더셀리스트, 데이터행리스트) 튜플을 순서대로 뽑는다."""
    tables, rows, header = [], [], None
    for line in text.split("\n"):
        s = line.strip()
        if s.startswith("|") and s.endswith("|"):
            cells = [c.strip() for c in s[1:-1].split("|")]
            if set("".join(cells)) <= set("-: "):  # 구분선
                continue
            if header is None:
                header = cells
            else:
                rows.append(cells)
        else:
            if header is not None:
                tables.append((header, rows))
            header, rows = None, []
    if header is not None:
        tables.append((header, rows))
    return tables


def clean(v):
    """마크다운 강조·링크 표기를 제거한다."""
    v = re.sub(r"\*\*(.+?)\*\*", r"\1", v)
    v = re.sub(r"\*(.+?)\*", r"\1", v)
    v = re.sub(r"`(.+?)`", r"\1", v)
    v = v.replace("⚠️", "").replace("  ", " ")
    return v.strip()


def week_key(w):
    """W00a → (0,0), W00b → (0,1), W07 → (7,0) 로 정렬 가능한 키."""
    m = re.match(r"W(\d+)([ab])?$", w)
    if not m:
        return (999, 0)
    return (int(m.group(1)), {"a": 0, "b": 1, None: 0}[m.group(2)])


def style_header(ws, row, ncols, fill=GREEN, size=10):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = PatternFill("solid", fgColor=fill)
        cell.font = Font(bold=True, color="FFFFFF", size=size)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER


def date_row(ws, weeks, col_offset, row, label="시작일"):
    """주차 컬럼 위에 시작 날짜 행을 깐다."""
    c0 = ws.cell(row=row, column=1, value=label)
    c0.font = Font(bold=True, size=9, color=GREEN_DK)
    c0.alignment = Alignment(horizontal="center", vertical="center")
    c0.border = BORDER
    for i, w in enumerate(weeks):
        cell = ws.cell(row=row, column=col_offset + i + 1, value=WEEK_START.get(w, ""))
        cell.font = Font(size=8, color=GREY_TX)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.fill = PatternFill("solid", fgColor=GREEN_LT)
        cell.border = BORDER


def mark_gates(ws, weeks, col_offset, first_row, last_row):
    """게이트 검수 주·내부 목표 주에 세로줄을 긋는다."""
    for w, side in [(g, GATE_SIDE) for g in GATES] + [(t, TGT_SIDE) for t in TARGETS]:
        if w not in weeks:
            continue
        c = col_offset + weeks.index(w) + 1
        for r in range(first_row, last_row + 1):
            cell = ws.cell(row=r, column=c)
            b = cell.border
            cell.border = Border(left=side, right=side, top=b.top, bottom=b.bottom)
        ws.cell(row=first_row, column=c).fill = PatternFill(
            "solid", fgColor=RED if side is GATE_SIDE else GREEN_DK)


def autosize(ws, widths):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


# ─────────────────────────── 시트 빌더 ───────────────────────────

def sheet_overview(wb, meta, gates, phases):
    ws = wb.create_sheet("개요")
    ws["A1"] = meta["title"]
    ws["A1"].font = Font(bold=True, size=15, color=GREEN_DK)
    ws["A2"] = f"문서 {meta.get('version','')} · 최종 {meta.get('date','')} · 자동 생성(마크다운 정본)"
    ws["A2"].font = Font(size=9, color=GREY_TX)

    r = 4
    ws.cell(row=r, column=1, value="일정 요약").font = Font(bold=True, size=12, color=GREEN_DK)
    r += 1
    ws.cell(row=r, column=1, value="항목")
    ws.cell(row=r, column=2, value="내용")
    style_header(ws, r, 2)
    r += 1
    for k, v in meta["summary"]:
        ws.cell(row=r, column=1, value=k).font = Font(bold=True)
        ws.cell(row=r, column=2, value=v).alignment = Alignment(wrap_text=True, vertical="top")
        for c in (1, 2):
            ws.cell(row=r, column=c).border = BORDER
        r += 1

    if phases:
        r += 1
        ws.cell(row=r, column=1, value="단계 구성").font = Font(bold=True, size=12, color=GREEN_DK)
        r += 1
        hdr = ["단계", "공식 구간", "내부 목표 구간"]
        for i, h in enumerate(hdr, start=1):
            ws.cell(row=r, column=i, value=h)
        style_header(ws, r, len(hdr))
        r += 1
        for row in phases:
            for i, v in enumerate(row[:3], start=1):
                cell = ws.cell(row=r, column=i, value=clean(v))
                cell.border = BORDER
                cell.alignment = Alignment(wrap_text=True, vertical="top")
            r += 1

    if gates:
        r += 1
        ws.cell(row=r, column=1, value="게이트 (통과 기준은 PRD §4-1 정본)").font = Font(
            bold=True, size=12, color=GREEN_DK)
        r += 1
        hdr = ["게이트", "공식(보고)", "내부 목표", "내부 실질 데드라인", "확보 버퍼"]
        for i, h in enumerate(hdr, start=1):
            ws.cell(row=r, column=i, value=h)
        style_header(ws, r, len(hdr))
        r += 1
        for row in gates:
            for i, v in enumerate(row[:5], start=1):
                cell = ws.cell(row=r, column=i, value=clean(v))
                cell.border = BORDER
                cell.alignment = Alignment(wrap_text=True, vertical="top")
            r += 1

    autosize(ws, [22, 40, 22, 30, 16])
    return ws


def sheet_weeks(wb, header, rows):
    ws = wb.create_sheet("주차체계")
    hdr = [clean(h) for h in header]
    for i, h in enumerate(hdr, start=1):
        ws.cell(row=1, column=i, value=h)
    style_header(ws, 1, len(hdr))
    ws.freeze_panes = "A2"

    for r, row in enumerate(rows, start=2):
        stage = clean(row[5]) if len(row) > 5 else ""
        for i, v in enumerate(row, start=1):
            cell = ws.cell(row=r, column=i, value=clean(v))
            cell.border = BORDER
            cell.alignment = Alignment(wrap_text=(i == len(row)), vertical="center")
            if i in (3, 5, 6):
                cell.alignment = Alignment(horizontal="center", vertical="center")
        # 단계 색
        if stage in STAGE_FILL:
            ws.cell(row=r, column=6).fill = PatternFill("solid", fgColor=STAGE_FILL[stage])
        # 저가용 주 강조
        if "⚠️" in row[2]:
            ws.cell(row=r, column=3).fill = PatternFill("solid", fgColor=AMBER)
            ws.cell(row=r, column=3).font = Font(bold=True, color=RED)
        # 게이트/오픈 행 강조
        note = clean(row[-1])
        if "Gate" in note or "오픈" in note:
            ws.cell(row=r, column=1).font = Font(bold=True, color=GREEN_DK)

    autosize(ws, [10, 18, 8, 26, 12, 8, 60])
    return ws


def sheet_resource(wb, blocks):
    """blocks: [(헤더셀, 행들)] — 주차 구간별 리소스 표를 하나로 합친다."""
    ws = wb.create_sheet("리소스판")
    weeks, table = [], {}
    for header, rows in blocks:
        ws_weeks = [clean(h) for h in header[1:]]
        for w in ws_weeks:
            if w not in weeks:
                weeks.append(w)
        for row in rows:
            name = clean(row[0])
            table.setdefault(name, {})
            for w, v in zip(ws_weeks, row[1:]):
                table[name][w] = clean(v)
    weeks.sort(key=week_key)

    date_row(ws, weeks, 1, 1)
    ws.cell(row=2, column=1, value="멤버 (포지션)")
    for i, w in enumerate(weeks, start=2):
        ws.cell(row=2, column=i, value=w)
    style_header(ws, 2, len(weeks) + 1, size=9)
    ws.freeze_panes = "B3"

    order = [n for n in table if "가용일" in n] + [n for n in table if "가용일" not in n]
    for r, name in enumerate(order, start=3):
        c0 = ws.cell(row=r, column=1, value=name)
        c0.border = BORDER
        c0.font = Font(bold=True, size=9)
        c0.alignment = Alignment(vertical="center")
        is_avail = "가용일" in name
        for i, w in enumerate(weeks, start=2):
            v = table[name].get(w, "")
            cell = ws.cell(row=r, column=i)
            cell.border = BORDER
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.font = Font(size=9)
            if is_avail:
                cell.value = v
                if v and v.strip() not in ("5", ""):
                    cell.fill = PatternFill("solid", fgColor=AMBER)
                    cell.font = Font(size=9, bold=True, color=RED)
                continue
            try:
                n = int(re.sub(r"[^0-9]", "", v)) if v else None
            except ValueError:
                n = None
            if n is None:
                continue
            cell.value = n
            cell.number_format = '0"%"'
            fill = RATE_FILL.get(n)
            if fill and n > 0:
                cell.fill = PatternFill("solid", fgColor=fill)
                cell.font = Font(size=9, bold=(n == 100),
                                 color="FFFFFF" if n >= 75 else "000000")
            elif n == 0:
                cell.fill = PatternFill("solid", fgColor=GREY)
                cell.font = Font(size=9, color=GREY_TX)

    mark_gates(ws, weeks, 1, 1, len(order) + 2)
    autosize(ws, [22] + [6] * len(weeks))
    ws.cell(row=len(order) + 4, column=1,
            value="100=풀투입 · 회색 0=미투입 · 가용일 행의 노랑=공휴일 주(3~4일) · "
                  f"빨간 세로줄={'·'.join(GATES)} 게이트 검수 · 초록 점선={'·'.join(TARGETS)} 내부 목표"
            ).font = Font(size=9, color=GREY_TX)
    return ws


def sheet_gantt(wb, feature_tables, ops_table, weeks, phases=None):
    ws = wb.create_sheet("WBS_간트")
    base = ["Epic (REQ)", "Story (FS)", "기능명", "스프린트", "리드", "전제·의존"]
    date_row(ws, weeks, len(base), 1)
    for i, h in enumerate(base, start=1):
        ws.cell(row=2, column=i, value=h)
    for i, w in enumerate(weeks, start=len(base) + 1):
        ws.cell(row=2, column=i, value=w)
    style_header(ws, 2, len(base) + len(weeks), size=9)
    ws.freeze_panes = f"{get_column_letter(len(base)+1)}3"

    widx = {w: i for i, w in enumerate(weeks)}
    r = 3

    def span_cells(span):
        m = re.findall(r"W\d+[ab]?", span or "")
        if not m or m[0] not in widx or m[-1] not in widx:
            return []
        return list(range(widx[m[0]], widx[m[-1]] + 1))

    # ── 단계 개요: 공식 구간(진한) 위에 내부 목표 구간(연한)을 겹쳐 그린다
    if phases:
        cell = ws.cell(row=r, column=1, value="단계 개요 — 진한색=공식(보고) 구간 · 연한색=내부 목표 선행 구간")
        cell.font = Font(bold=True, size=10, color="FFFFFF")
        for c in range(1, len(base) + len(weeks) + 1):
            ws.cell(row=r, column=c).fill = PatternFill("solid", fgColor=GREEN_DK)
        r += 1
        pcolors = [("9DC3E6", "DEEAF6"), ("B4A7D6", "E4DFEC"),
                   ("F4A6B0", "FCE4E6"), ("EA9AA6", "F9DDE1")]
        for i, prow in enumerate(phases):
            name, official, internal = (clean(prow[0]), clean(prow[1]), clean(prow[2]))
            dk, lt = pcolors[i % len(pcolors)]
            for j, v in enumerate([name, "", f"공식 {official}", "", "", f"내부 목표 {internal}"], start=1):
                c = ws.cell(row=r, column=j, value=v)
                c.border = BORDER
                c.font = Font(size=9, bold=(j == 1))
                c.alignment = Alignment(wrap_text=(j == 6), vertical="center")
            for k in span_cells(internal):
                cc = ws.cell(row=r, column=len(base) + 1 + k)
                cc.fill = PatternFill("solid", fgColor=lt)
                cc.border = BORDER
            for k in span_cells(official):
                cc = ws.cell(row=r, column=len(base) + 1 + k)
                cc.fill = PatternFill("solid", fgColor=dk)
                cc.border = BORDER
            r += 1
        r += 1

    def bar(row_i, span, color):
        m = re.findall(r"W\d+[ab]?", span or "")
        if not m:
            return
        s, e = m[0], m[-1]
        if s not in widx or e not in widx:
            return
        for k in range(widx[s], widx[e] + 1):
            cell = ws.cell(row=row_i, column=len(base) + 1 + k)
            cell.fill = PatternFill("solid", fgColor=color)
            cell.border = BORDER

    def write(cells, phase_color):
        nonlocal r
        # cells: Epic, Story, 기능명, 주차, 스프린트, 리드, 전제
        vals = [cells[0], cells[1], cells[2], cells[4], cells[5], cells[6]]
        for i, v in enumerate(vals, start=1):
            cell = ws.cell(row=r, column=i, value=clean(v))
            cell.border = BORDER
            cell.font = Font(size=9)
            cell.alignment = Alignment(wrap_text=(i in (3, 6)), vertical="top")
        if "🔴" in cells[6]:
            ws.cell(row=r, column=6).fill = PatternFill("solid", fgColor=AMBER)
        bar(r, cells[3], phase_color)
        r += 1

    for title, (header, rows), color in feature_tables:
        cell = ws.cell(row=r, column=1, value=title)
        cell.font = Font(bold=True, size=10, color="FFFFFF")
        for c in range(1, len(base) + len(weeks) + 1):
            ws.cell(row=r, column=c).fill = PatternFill("solid", fgColor=GREEN_DK)
        r += 1
        for row in rows:
            if len(row) >= 7:
                write(row, color)

    if ops_table:
        header, rows = ops_table
        cell = ws.cell(row=r, column=1, value="공정·운영 태스크 (OPS)")
        cell.font = Font(bold=True, size=10, color="FFFFFF")
        for c in range(1, len(base) + len(weeks) + 1):
            ws.cell(row=r, column=c).fill = PatternFill("solid", fgColor=GREEN_DK)
        r += 1
        for row in rows:
            if len(row) < 6:
                continue
            vals = [row[0], "", row[1], row[3], row[4], row[5]]
            for i, v in enumerate(vals, start=1):
                cell = ws.cell(row=r, column=i, value=clean(v))
                cell.border = BORDER
                cell.font = Font(size=9)
                cell.alignment = Alignment(wrap_text=(i in (3, 6)), vertical="top")
            bar(r, row[2], "F4B183")
            r += 1

    mark_gates(ws, weeks, len(base), 1, r - 1)
    autosize(ws, [12, 12, 34, 12, 12, 40] + [3.2] * len(weeks))
    ws.cell(row=r + 1, column=1,
            value=f"빨간 세로줄={'·'.join(GATES)} 게이트 검수 · 초록 점선={'·'.join(TARGETS)} 내부 목표 "
                  "· 태스크 바=공식(보고) 주차 · 앰버=전제 건 · 버퍼 주(W13·W22)에 태스크가 없는 것은 "
                  "내부 선행으로 다음 단계가 이미 착수되기 때문(단계 개요 참조)").font = Font(size=9, color=GREY_TX)
    return ws


def sheet_open(wb, header, rows):
    ws = wb.create_sheet("미결")
    hdr = [clean(h) for h in header]
    for i, h in enumerate(hdr, start=1):
        ws.cell(row=1, column=i, value=h)
    style_header(ws, 1, len(hdr))
    for r, row in enumerate(rows, start=2):
        for i, v in enumerate(row, start=1):
            cell = ws.cell(row=r, column=i, value=clean(v))
            cell.border = BORDER
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            cell.font = Font(size=10)
    autosize(ws, [8, 90, 24])
    return ws


# ─────────────────────────── 메인 ───────────────────────────

def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("project", help="docs/[프로젝트명] 경로")
    ap.add_argument("--out", help="출력 xlsx 경로 (기본: wbs/ 하위 동일 파일명)")
    ap.add_argument("--update", action="store_true",
                    help="기존 xlsx를 열어 셀 '값'만 갱신 — 폰트·색·열너비 등 수작업 서식 보존")
    ap.add_argument("--recolor", action="store_true",
                    help="--update 시 투입률 셀 색·간트 바도 값에 맞춰 다시 칠한다")
    args = ap.parse_args()

    proj = Path(args.project)
    wbs_files = sorted((proj / "wbs").glob("WBS_*.md"))
    if not wbs_files:
        sys.exit(f"❌ WBS 마크다운을 찾지 못함: {proj}/wbs/WBS_*.md")
    src = wbs_files[0]
    text = src.read_text(encoding="utf-8")
    tables = md_tables(text)

    weeks_tbl = gates_tbl = phases_tbl = open_tbl = ops_tbl = None
    res_blocks, feat_tbls = [], []
    for header, rows in tables:
        h0 = clean(header[0])
        if h0 == "주차" and weeks_tbl is None:
            weeks_tbl = (header, rows)
        elif h0 == "게이트":
            gates_tbl = rows
        elif h0 == "단계" and len(header) >= 4 and "내부 목표" in "".join(header):
            phases_tbl = rows
        elif h0 == "멤버 (포지션)":
            res_blocks.append((header, rows))
        elif h0.startswith("Epic"):
            feat_tbls.append((header, rows))
        elif h0 == "Task ID":
            ops_tbl = (header, rows)
        elif h0 == "#" and "항목" in "".join(clean(c) for c in header):
            open_tbl = (header, rows)

    if not weeks_tbl:
        sys.exit("❌ 주차 체계 표를 찾지 못함")

    weeks = [clean(r[0]) for r in weeks_tbl[1]]

    global GATES, TARGETS, WEEK_START
    WEEK_START = {clean(row[0]): clean(row[1]).split("~")[0].strip()
                  for row in weeks_tbl[1] if len(row) > 1}
    GATES = [clean(r[0]) for r in weeks_tbl[1] if "Gate" in clean(r[-1]) and "검수" in clean(r[-1])]
    TARGETS = [clean(r[0]) for r in weeks_tbl[1] if "내부 Gate" in clean(r[-1])]

    # 메타
    title = text.split("\n")[0].lstrip("# ").strip()
    ver = re.search(r"문서 버전:\s*(\S+)", text)
    date = re.search(r"최종\s*(\d{4}-\d{2}-\d{2})", text)
    prem = re.search(r"- 일정 전제:\s*(.+)", text)
    summary = [("일정 전제", clean(prem.group(1)) if prem else "")]
    for label, pat in [
        ("주차 범위", r"## 1\. 주차 체계 \((.+?)\)"),
        ("실가용", r"(W01~W30 총 \d+일[^.]*)"),
    ]:
        m = re.search(pat, text)
        if m:
            summary.append((label, clean(m.group(1))))
    summary.append(("시트 구성", "개요 / 주차체계 / 리소스판 / WBS_간트 / 미결"))
    summary.append(("주의", f"자동 생성물 — 수정은 {src.name} 에서 하고 재실행"))

    if args.update:
        out = Path(args.out) if args.out else src.with_suffix(".xlsx")
        if not out.exists():
            sys.exit(f"❌ --update 대상 파일이 없음: {out}")
        res = {}
        for header, rows in tables:
            if clean(header[0]) == "멤버 (포지션)":
                wk = [clean(h) for h in header[1:]]
                for row in rows:
                    res.setdefault(clean(row[0]), {}).update(
                        {w: clean(v) for w, v in zip(wk, row[1:])})
        feat = {}
        for header, rows in tables:
            h0 = clean(header[0])
            if h0.startswith("Epic"):
                for row in rows:
                    feat[(clean(row[0]), clean(row[1]))] = (
                        [clean(row[2]), clean(row[4]), clean(row[5]), clean(row[6])], clean(row[3]))
            elif h0 == "Task ID":
                for row in rows:
                    feat[(clean(row[0]), "")] = (
                        [clean(row[1]), clean(row[3]), clean(row[4]), clean(row[5])], clean(row[2]))
        rep = update_in_place(out, {"weeks": weeks_tbl[1], "res": res, "feat": feat},
                              weeks, recolor=args.recolor)
        print(f"✅ 값 갱신(서식 보존): {out}")
        print(f"   갱신 {rep['updated']}행" + ("  · 색 재적용" if args.recolor else "  · 색 유지"))
        for k in rep["missing"]:
            print(f"   ⚠️ md에는 있으나 엑셀에 행이 없음 — {k}")
        for k in rep["unmatched"][:10]:
            print(f"   ℹ️ 엑셀에만 있는 행(그대로 둠) — {k}")
        return

    wb = Workbook()
    wb.remove(wb.active)
    sheet_overview(wb, {
        "title": title,
        "version": ver.group(1) if ver else "",
        "date": date.group(1) if date else "",
        "summary": summary,
    }, gates_tbl, phases_tbl)
    sheet_weeks(wb, *weeks_tbl)
    if res_blocks:
        sheet_resource(wb, res_blocks)
    if feat_tbls:
        colors = ["9DC3E6", "B4A7D6", "F4A6B0", "EA9AA6", "A9D08E"]
        labeled = []
        titles = re.findall(r"### (3-\d\..+)", text)
        for i, t in enumerate(feat_tbls):
            label = clean(titles[i]) if i < len(titles) else f"구간 {i+1}"
            labeled.append((label, t, colors[i % len(colors)]))
        sheet_gantt(wb, labeled, ops_tbl, weeks, phases_tbl)
    if open_tbl:
        sheet_open(wb, *open_tbl)

    out = Path(args.out) if args.out else src.with_suffix(".xlsx")
    wb.save(out)
    print(f"✅ 생성: {out}")
    print(f"   시트: {', '.join(wb.sheetnames)}")


def update_in_place(path, tables_, weeks, recolor=False):
    """기존 통합문서의 값만 갱신한다(서식 보존). 시트·행 구조는 건드리지 않는다."""
    from openpyxl import load_workbook as _lw
    wb = _lw(path)
    report = {"updated": 0, "unmatched": [], "missing": []}

    def find_row(ws, text, col=1, upto=None):
        for r in range(1, (upto or ws.max_row) + 1):
            if str(ws.cell(row=r, column=col).value or "").strip() == text:
                return r
        return None

    # ── 주차체계
    if "주차체계" in wb.sheetnames:
        ws = wb["주차체계"]
        hdr = find_row(ws, "주차")
        if hdr:
            byweek = {clean(row[0]): row for row in tables_["weeks"]}
            for r in range(hdr + 1, ws.max_row + 1):
                key = str(ws.cell(row=r, column=1).value or "").strip()
                if key in byweek:
                    for i, v in enumerate(byweek[key][1:], start=2):
                        ws.cell(row=r, column=i).value = clean(v)
                    report["updated"] += 1
                elif key:
                    report["unmatched"].append(f"주차체계 {r}행 '{key}'")

    # ── 리소스판
    if "리소스판" in wb.sheetnames:
        ws = wb["리소스판"]
        hdr = find_row(ws, "멤버 (포지션)")
        if hdr:
            cols = {}
            for c in range(2, ws.max_column + 1):
                w = str(ws.cell(row=hdr, column=c).value or "").strip()
                if w:
                    cols[w] = c
            for r in range(hdr + 1, ws.max_row + 1):
                name = str(ws.cell(row=r, column=1).value or "").strip()
                if not name or name not in tables_["res"]:
                    if name and "=" not in name:
                        report["unmatched"].append(f"리소스판 {r}행 '{name}'")
                    continue
                for w, v in tables_["res"][name].items():
                    if w not in cols:
                        continue
                    cell = ws.cell(row=r, column=cols[w])
                    if "가용일" in name:
                        cell.value = v
                    else:
                        try:
                            n = int(re.sub(r"[^0-9]", "", v)) if v else None
                        except ValueError:
                            n = None
                        cell.value = n
                        if recolor and n is not None:
                            fill = RATE_FILL.get(n)
                            if fill and n > 0:
                                cell.fill = PatternFill("solid", fgColor=fill)
                            elif n == 0:
                                cell.fill = PatternFill("solid", fgColor=GREY)
                report["updated"] += 1
            for name in tables_["res"]:
                if not find_row(ws, name):
                    report["missing"].append(f"리소스판 '{name}' (엑셀에 행 없음)")

    # ── WBS_간트
    if "WBS_간트" in wb.sheetnames:
        ws = wb["WBS_간트"]
        hdr = find_row(ws, "Epic (REQ)")
        if hdr:
            wcol = {}
            for c in range(1, ws.max_column + 1):
                w = str(ws.cell(row=hdr, column=c).value or "").strip()
                if re.match(r"W\d+[ab]?$", w):
                    wcol[w] = c
            seen = set()
            for r in range(hdr + 1, ws.max_row + 1):
                a = str(ws.cell(row=r, column=1).value or "").strip()
                b = str(ws.cell(row=r, column=2).value or "").strip()
                key = (a, b)
                if key not in tables_["feat"]:
                    continue
                vals, span = tables_["feat"][key]
                for i, v in enumerate(vals, start=3):
                    ws.cell(row=r, column=i).value = v
                seen.add(key)
                report["updated"] += 1
                if recolor and span:
                    m = re.findall(r"W\d+[ab]?", span)
                    if m and m[0] in wcol and m[-1] in wcol:
                        for c in range(wcol[m[0]], wcol[m[-1]] + 1):
                            ws.cell(row=r, column=c).fill = PatternFill("solid", fgColor="9DC3E6")
            for key in tables_["feat"]:
                if key not in seen:
                    report["missing"].append(f"간트 {key[0]} {key[1]} (엑셀에 행 없음)")

    wb.save(path)
    return report


if __name__ == "__main__":
    main()
