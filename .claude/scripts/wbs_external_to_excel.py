#!/usr/bin/env python3
"""외부 공유용 WBS 마크다운 → 엑셀 변환 (개요·주차체계·WBS 간트·게이트·선행제공물·제외)

사용법:
    python3 .claude/scripts/wbs_external_to_excel.py docs/[프로젝트명]
    python3 .claude/scripts/wbs_external_to_excel.py docs/[프로젝트명] --out 경로.xlsx

`wbs/WBS_외부공유용_*.md`를 읽어 시트 6개로 변환한다.
내부 WBS 변환기(wbs_to_excel.py)와 달리 **리소스판·Jira 운영 규칙을 담지 않는다** —
발주사·수행사 밖으로 나가는 문서이므로 인력 투입률 등 내부 정보를 싣지 않는다.

자동 생성물이므로 엑셀을 직접 고치지 말고 마크다운을 고친 뒤 재실행한다.
"""
import argparse
import re
import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# 라이프플래닛 그린 계열(.claude/assets/pdf 팔레트와 동일 계통)
GREEN = "00A862"
GREEN_DK = "007A47"
GREEN_LT = "E6F6EF"
GREY = "F2F2F2"
AMBER = "FFF4CE"
RED = "C00000"

THIN = Side(style="thin", color="D9D9D9")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
GATE_SIDE = Side(style="medium", color=RED)

# 단계별 간트 바 색
PHASE_FILL = {
    "P1": "1F3B66",
    "P2": "2E5A88",
    "P3-1": "2F8F83",
    "P3-2": "00A862",
    "전구간": "9BA7B5",
}
PHASE_LT = {
    "P1": "D9E4F1",
    "P2": "DEE8F3",
    "P3-1": "D9EEEA",
    "P3-2": "D9F2E6",
    "전구간": "ECEFF3",
}


def md_tables(text):
    """마크다운에서 (헤더셀, 데이터행들) 튜플을 순서대로 뽑는다."""
    tables, rows, header = [], [], None
    for line in text.split("\n"):
        s = line.strip()
        if s.startswith("|") and s.endswith("|"):
            cells = [c.strip() for c in s.strip("|").split("|")]
            if re.fullmatch(r"[\s:\-]+", "".join(cells)):
                continue                      # 구분선
            if header is None:
                header = cells
            else:
                rows.append(cells)
        else:
            if header:
                tables.append((header, rows))
            header, rows = None, []
    if header:
        tables.append((header, rows))
    return tables


def clean(v):
    """엑셀에 넣기 전 마크다운 강조·각주 표기를 걷어낸다."""
    v = re.sub(r"\*\*(.+?)\*\*", r"\1", v or "")
    v = re.sub(r"\*(.+?)\*", r"\1", v)
    v = v.replace("~~", "").replace("`", "")
    return v.strip()


def week_cols(n_start=1, n_end=32):
    return [f"W{i:02d}" for i in range(n_start, n_end + 1)]


def style_header(ws, row, ncols, fill=GREEN, size=10):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = PatternFill("solid", fgColor=fill)
        cell.font = Font(bold=True, color="FFFFFF", size=size)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER


def autosize(ws, widths):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def title_block(ws, title, note=None):
    ws["A1"] = title
    ws["A1"].font = Font(bold=True, size=14, color=GREEN_DK)
    r = 2
    if note:
        ws.cell(row=2, column=1, value=note).font = Font(size=9, color="666666")
        ws.cell(row=2, column=1).alignment = Alignment(wrap_text=True, vertical="top")
        r = 3
    return r + 1


# ─────────────────────────── 시트 빌더 ───────────────────────────

def sheet_overview(wb, meta_tbl, phases, gates):
    ws = wb.create_sheet("개요")
    r = title_block(ws, "AI 세일즈봇 구축 WBS — 외부 공유용",
                    "요구사항(REQ)과 기능(FS)이 어느 주차에 배치되는지, 그 착수에 필요한 발주사 선행 제공물이 무엇인지 정리한 문서입니다. "
                    "기능의 처리 규칙은 기능명세서, 인수 조건은 작업명세서(SOW)를 따릅니다.")
    ws.cell(row=r, column=1, value="문서 정보").font = Font(bold=True, size=11)
    r += 1
    style_header(ws, r, 2)
    ws.cell(row=r, column=1, value="구분"); ws.cell(row=r, column=2, value="내용")
    r += 1
    for row in meta_tbl:
        ws.cell(row=r, column=1, value=clean(row[0])).border = BORDER
        ws.cell(row=r, column=2, value=clean(row[1])).border = BORDER
        ws.cell(row=r, column=2).alignment = Alignment(wrap_text=True, vertical="top")
        r += 1

    r += 1
    ws.cell(row=r, column=1, value="단계 구성").font = Font(bold=True, size=11)
    r += 1
    style_header(ws, r, 4)
    for i, h in enumerate(["단계", "주차", "스프린트", "완료 모습"], start=1):
        ws.cell(row=r, column=i, value=h)
    r += 1
    for p in phases:
        for i, v in enumerate(p, start=1):
            c = ws.cell(row=r, column=i, value=v)
            c.border = BORDER
            c.alignment = Alignment(wrap_text=True, vertical="top")
        ws.cell(row=r, column=1).fill = PatternFill("solid", fgColor=PHASE_LT.get(p[0].split()[0], GREY))
        r += 1

    r += 1
    ws.cell(row=r, column=1, value="검수 게이트").font = Font(bold=True, size=11)
    r += 1
    style_header(ws, r, 3)
    for i, h in enumerate(["게이트", "시점", "통과 기준"], start=1):
        ws.cell(row=r, column=i, value=h)
    r += 1
    for g in gates:
        for i, v in enumerate(g[:3], start=1):
            c = ws.cell(row=r, column=i, value=clean(v))
            c.border = BORDER
            c.alignment = Alignment(wrap_text=True, vertical="top")
        r += 1
    autosize(ws, [22, 30, 20, 70])
    return ws


def sheet_weeks(wb, header, rows):
    ws = wb.create_sheet("주차체계")
    r = title_block(ws, "주차 체계 (W01~W32)",
                    "기간은 킥오프 가정 기준이며 절대일은 계약 시 확정합니다. 공휴일이 걸린 주는 용량을 줄여 계획했습니다.")
    style_header(ws, r, len(header))
    for i, h in enumerate(header, start=1):
        ws.cell(row=r, column=i, value=h)
    hdr = r
    r += 1
    for row in rows:
        stage = clean(row[3]) if len(row) > 3 else ""
        gate = "게이트" in clean(row[-1])
        for i, v in enumerate(row, start=1):
            c = ws.cell(row=r, column=i, value=clean(v))
            c.border = BORDER
            c.alignment = Alignment(wrap_text=True, vertical="center",
                                    horizontal="center" if i in (1, 3, 4) else "left")
            if i == 4 and stage in PHASE_LT:
                c.fill = PatternFill("solid", fgColor=PHASE_LT[stage])
        if gate:
            for i in range(1, len(row) + 1):
                ws.cell(row=r, column=i).font = Font(bold=True, color=RED)
        r += 1
    ws.freeze_panes = ws.cell(row=hdr + 1, column=1)
    autosize(ws, [10, 18, 12, 10, 52])
    return ws


def sheet_gantt(wb, sections, weeks):
    ws = wb.create_sheet("WBS_간트")
    r = title_block(ws, "WBS — Epic(REQ) × Story(FS) × 주차",
                    "막대는 계획 시작~종료 주입니다. '발주사 선행'이 채워진 항목은 해당 제공물이 없으면 착수할 수 없습니다.")
    base = ["Epic (REQ)", "Story (FS)", "기능명", "스프린트", "산출물", "발주사 선행"]
    ncol = len(base) + len(weeks)
    style_header(ws, r, ncol)
    for i, h in enumerate(base, start=1):
        ws.cell(row=r, column=i, value=h)
    for j, w in enumerate(weeks, start=len(base) + 1):
        ws.cell(row=r, column=j, value=w)
        ws.column_dimensions[get_column_letter(j)].width = 3.4
    hdr = r
    r += 1

    for label, phase, rows in sections:
        ws.cell(row=r, column=1, value=label).font = Font(bold=True, size=10, color="FFFFFF")
        for c in range(1, ncol + 1):
            ws.cell(row=r, column=c).fill = PatternFill("solid", fgColor=PHASE_FILL.get(phase, GREEN_DK))
            ws.cell(row=r, column=c).border = BORDER
        r += 1
        for row in rows:
            cells = [clean(x) for x in row]
            span = ""
            # 주차 칸 위치: 헤더에 '주차'가 있는 열
            for i, v in enumerate(cells):
                if re.match(r"W\d{2}(~W\d{2})?$", v):
                    span = v
                    cells.pop(i)
                    break
            for i, v in enumerate(cells[:len(base)], start=1):
                c = ws.cell(row=r, column=i, value=v)
                c.border = BORDER
                c.alignment = Alignment(wrap_text=True, vertical="top",
                                        horizontal="center" if i <= 2 else "left")
            m = re.findall(r"W(\d{2})", span)
            if m:
                a, b = int(m[0]), int(m[-1])
                for wk in range(a, b + 1):
                    col = len(base) + wk
                    if col <= ncol:
                        ws.cell(row=r, column=col).fill = PatternFill("solid", fgColor=PHASE_FILL.get(phase, GREEN))
            for j in range(len(base) + 1, ncol + 1):
                ws.cell(row=r, column=j).border = BORDER
            r += 1
    ws.freeze_panes = ws.cell(row=hdr + 1, column=len(base) + 1)
    autosize(ws, [11, 10, 34, 12, 26, 30])
    return ws


def sheet_simple(wb, name, title, note, header, rows, widths):
    ws = wb.create_sheet(name)
    r = title_block(ws, title, note)
    style_header(ws, r, len(header))
    for i, h in enumerate(header, start=1):
        ws.cell(row=r, column=i, value=h)
    hdr = r
    r += 1
    for row in rows:
        for i, v in enumerate(row, start=1):
            c = ws.cell(row=r, column=i, value=clean(v))
            c.border = BORDER
            c.alignment = Alignment(wrap_text=True, vertical="top")
        r += 1
    ws.freeze_panes = ws.cell(row=hdr + 1, column=1)
    autosize(ws, widths)
    return ws


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("project", help="docs/[프로젝트명]")
    ap.add_argument("--out")
    args = ap.parse_args()

    proj = Path(args.project)
    cands = sorted(proj.glob("wbs/WBS_외부공유용*.md"))
    if not cands:
        print(f"❌ {proj}/wbs/WBS_외부공유용*.md 없음", file=sys.stderr)
        sys.exit(1)
    src = cands[-1]
    text = src.read_text(encoding="utf-8")
    tables = md_tables(text)

    def pick(*keys):
        """헤더 첫 칸들이 keys와 맞는 첫 표를 돌려준다."""
        for header, rows in tables:
            if all(k in " ".join(header) for k in keys):
                return header, rows
        return None, []

    meta_h, meta = pick("구분", "내용")
    weeks_h, weeks_rows = pick("주차", "스프린트", "단계")
    gate_h, gate_rows = pick("게이트", "시점", "통과 기준")
    pre_h, pre_rows = pick("필요 시점", "제공물")
    exc_h, exc_rows = pick("제외 항목", "대체")
    open_h, open_rows = pick("확정 주체")

    # 섹션별 간트 표 — 헤더에 'Epic (REQ)'가 든 표를 순서대로
    feat = [(h, r) for h, r in tables if "Epic" in " ".join(h)]
    labels = [
        ("P1 넛징·코어 (W01~W12 · SP-01~06)", "P1"),
        ("P2 추천·설득·Admin (W14~W21 · SP-07~10)", "P2"),
        ("P3-1 보험료·전환·운영 (W23~W26 · SP-11~12)", "P3-1"),
        ("전 구간 태스크", "전구간"),
    ]
    sections = [(labels[i][0], labels[i][1], rows) for i, (h, rows) in enumerate(feat) if i < len(labels)]

    # P3-2는 FS가 아닌 공정 표 — '구분/작업' 헤더
    p32_h, p32_rows = pick("구분", "작업", "주차")

    phases = [
        ["P1 넛징·코어", "W01~W12", "SP-01~06", "앞단 넛징으로 진입시켜 근거 상담까지"],
        ["P2 추천·설득·Admin", "W14~W21", "SP-07~10", "추천·비교·설득으로 웹 청약 연결/핸드오프까지"],
        ["P3-1 보험료·전환·운영", "W23~W26", "SP-11~12", "계정계 연동 정확 보험료·가입설계 심화"],
        ["P3-2 통합·안정화", "W27~W30", "SP-13~14", "통합 QA·부하·보안·기술이전"],
        ["오픈", "W31~W32", "—", "정식 오픈·안정화"],
    ]

    wb = Workbook()
    wb.remove(wb.active)
    sheet_overview(wb, meta, phases, gate_rows)
    if weeks_rows:
        sheet_weeks(wb, [clean(h) for h in weeks_h], weeks_rows)
    if sections:
        sheet_gantt(wb, sections, week_cols())
    if p32_rows:
        sheet_simple(wb, "P3-2_통합안정화", "P3-2 통합·안정화 (W27~W30)",
                     "기능 개발이 아니라 통합 검증·안정화·인도 구간입니다.",
                     [clean(h) for h in p32_h], p32_rows, [12, 34, 14, 12, 28, 30])
    if pre_rows:
        sheet_simple(wb, "선행제공물", "발주사 선행 제공물 · 필요 시점",
                     "아래 제공물이 해당 시점에 확보되어야 후속 작업이 계획대로 진행됩니다.",
                     [clean(h) for h in pre_h], pre_rows, [18, 52, 44])
    rows2 = list(exc_rows)
    if exc_rows:
        sheet_simple(wb, "제외범위", "본 사업 제외", None,
                     [clean(h) for h in exc_h], rows2, [46, 46])
    if open_rows:
        sheet_simple(wb, "확정필요", "확정이 필요한 사항", None,
                     [clean(h) for h in open_h], open_rows, [52, 20, 20])

    out = Path(args.out) if args.out else src.with_suffix(".xlsx")
    wb.save(out)
    print(f"✅ 생성: {out}")
    print(f"   시트: {', '.join(wb.sheetnames)}")


if __name__ == "__main__":
    main()
